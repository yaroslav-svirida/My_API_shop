import openpyxl
from django.http import Http404, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from myshop.models import Product, Basket, ProductInBasket
from myshop.serializers import BasketSerializer, ProductInBasketSerializer


class BasketView(APIView):
    def post(self, request):
        id = request.data.get("id")
        basket = Basket.objects.get(id=6)
        product = Product.objects.get(id=id)
        # product.amount_for_basket = request.data.get("amount_for_basket")
        # product.save()
        basket.products_to_basket.add(product)
        basket.save()
        user_basket_serialized = BasketSerializer(basket).data
        return Response(user_basket_serialized)

    def get(self, request):
        basket = Basket.objects.get(id=6)
        product_serialized = BasketSerializer(basket).data
        return Response(product_serialized)

    # def delete(self, request, pk):
    #     try:
    #         basket = Basket.objects.get(pk=pk)
    #     except Basket.DoesNotExist:
    #         raise Http404
    #     basket.delete()
    #     return Response(status=status.HTTP_204_NO_CONTENT)

#
# def put(self, request, pk):
#     user_id = request.user.id
#     basket = Basket.objects.get(user_id=user_id)
#     product = Product.objects.get(id=pk)
#     product.amount_for_basket = request.data.get("amount_for_basket")
#     product.save()
#     basket.products_to_basket.add(product)
#     basket.save()
#     user_basket_serialized = BasketSerializer(basket).data
#     return Response(user_basket_serialized)


class ProductInBasketView(APIView):
    def get(self, request):
        user_id = request.user.id
        basket= Basket.objects.get(user_id=user_id)

        basket_id = basket.id
        products = ProductInBasket.objects.filter(basket_id=basket_id)
        products_serialized = ProductInBasketSerializer(products,many=True).data
        return Response(products_serialized)

    def delete(self, request, pk):
        try:
            product = ProductInBasket.objects.get(product_id=pk)
        except ProductInBasket.DoesNotExist:
            raise Http404
        product.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    def post(self, request):
        product_id = request.data.get("id")
        product = Product.objects.get(id=product_id)
        basket = Basket.objects.get(user_id=request.user.id)
        basket_id = basket.id
        product_add = ProductInBasket.objects.create(
            basket_id=basket_id,
            product_id=product.id,
            product_name=product.name,
            product_article=product.article,
            product_price=product.price,
            product_amount=product.amount
        )

        product_serialized = ProductInBasketSerializer(product_add).data
        return Response(product_serialized)

    def put(self, request, pk):

        basket_id = 10
        product = ProductInBasket.objects.get(basket_id=basket_id, product_id=pk)
        serializer = ProductInBasketSerializer(product, data=request.data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ExportBasketInExcel(APIView):
    def get(self, request):
        user_id = request.user.id
        basket = Basket.objects.get(user_id=user_id)

        products_for_basket = ProductInBasket.objects.filter(basket_id=basket.id)
        print(products_for_basket)

        alignment = Alignment(horizontal='center',
                              vertical='center', )
        workbook = Workbook()
        ws = workbook.active
        fields = ['№ п/п', 'Артикул', 'Наименование', 'Количество', 'Цена за шт с НДС', 'Стоимость продукции']
        column = 1
        row = 1
        column_alf = ['A', 'B', 'C', 'D', 'E', 'F']
        column_weight = [10, 15, 90, 20, 20, 24]

        for field in fields:
            ws.column_dimensions[column_alf[column - 1]].width = column_weight[column - 1]
            cell = ws.cell(column=column, row=1, value=field)

            cell.alignment = alignment
            column += 1

        number_of_product = 1

        full_price_for_all_product = 0

        for product in products_for_basket:

            row += 1

            ws.cell(column=1, row=row, value=number_of_product).alignment = alignment
            ws.cell(column=2, row=row, value=product.product_article).alignment = alignment
            ws.cell(column=3, row=row, value=product.product_name)
            ws.cell(column=4, row=row, value=product.product_amount).alignment = alignment
            ws.cell(column=5, row=row, value=product.product_price).alignment = alignment
            full_price_for_one_product = product.product_price * product.product_amount
            ws.cell(column=6, row=row, value=full_price_for_one_product).alignment = alignment

            full_price_for_all_product += full_price_for_one_product

            number_of_product += 1

            if number_of_product - 1 == len(list(products_for_basket)):
                print('boom')
                ws.cell(column=5, row=row + 1, value='Итого').alignment = alignment
                ws.cell(column=6, row=row + 1, value=full_price_for_all_product).alignment = alignment

        response = HttpResponse(content_type='application/ms_excel')
        response['Content_Disposition'] = f'attachment; filename = Data.xlsx'
        workbook.save(response)
        return response
