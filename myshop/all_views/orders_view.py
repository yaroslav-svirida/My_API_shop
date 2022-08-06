from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from rest_framework.response import Response
from rest_framework.views import APIView

from myshop.models import Basket, Orders, Product
from myshop.serializers import OrderSerializer, BasketSerializer


def get(request):
    basket = Basket.objects.get(user_id=request.user.id)
    product_serialized = BasketSerializer(basket).data
    return Response(product_serialized)


class OrdersView(APIView):

    def post(self, request):
        user_id = request.user.id
        basket = Basket.objects.get(user_id=user_id)
        # products = basket.products_to_basket.all()
        #
        # order = Orders.objects.create(
        #     user=request.user,
        # )
        #
        # for prod in products:
        #     order.order_product.add(prod)
        # order.save()
        order_serialized = OrderSerializer(basket).data

        return Response(order_serialized)


# def delete(self, request, pk):
#     user_id = request.user.id
#     basket = Basket.objects.get(id=user_id)
#     product = Product.objects.get(id=pk)
#     product.to_basket = 0
#     product.save()
#     basket.products_to_basket.remove(product)
#     return Response(status=status.HTTP_204_NO_CONTENT)


class ExportOrderInExcel(APIView):
    def get(self, request, pk):
        user_id = request.user.id
        order = Orders.objects.get(id=pk)
        products = order.order_product.all()
        print(products)

        alignment = Alignment(horizontal='center',
                              vertical='center', )
        workbook = Workbook()
        ws = workbook.active
        fields = ['№ п/п', 'Артикул', 'Наименование', 'Количество', 'Цена за шт с НДС', 'Стоимость продукции']
        column = 1
        row = 1
        column_alf = ['A', 'B', 'C', 'D', 'E', 'F']
        column_weight = [10, 15, 90, 20, 20, 24]

        for field in fields:
            ws.column_dimensions[column_alf[column - 1]].width = column_weight[column - 1]
            cell = ws.cell(column=column, row=1, value=field)

            cell.alignment = alignment
            column += 1

        number_of_product = 1

        full_price_for_all_product = 0

        for product in products:
            row += 1

            ws.cell(column=1, row=row, value=number_of_product).alignment = alignment
            ws.cell(column=2, row=row, value=product.article).alignment = alignment
            ws.cell(column=3, row=row, value=product.name)
            ws.cell(column=4, row=row, value=product.amount).alignment = alignment
            ws.cell(column=5, row=row, value=product.price).alignment = alignment
            full_price_for_one_product = product.price * product.amount
            ws.cell(column=6, row=row, value=full_price_for_one_product).alignment = alignment

            full_price_for_all_product += full_price_for_one_product

            number_of_product += 1

            if number_of_product - 1 == len(list(products)):
                print('boom')
                ws.cell(column=5, row=row + 1, value='Итого').alignment = alignment
                ws.cell(column=6, row=row + 1, value=full_price_for_all_product).alignment = alignment

        response = HttpResponse(content_type='application/ms_excel')
        response['Content_Disposition'] = f'attachment; filename = Data.xlsx'
        workbook.save(response)
        return response
