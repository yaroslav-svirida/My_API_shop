import openpyxl
from django.http import Http404, HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import Product, User, Basket
from myshop.serializers import ProductSerializer, UserSerializer, ProductBaskerSerializer, BasketSerializer


class BasketView(APIView):
    '''Create basket if not exists'''

    def post(self, request):

        user_id = request.user.id
        if Basket.objects.get(id=user_id):
            basket = Basket.objects.get(id=user_id)
            products_for_basket = Product.objects.filter(to_basket=True)



            for prod in products_for_basket:
                basket.products_to_basket.add(prod)
            basket.save()
            user_basket_serialized = BasketSerializer(basket).data
            return Response(user_basket_serialized)

        else:
            basket = Basket.objects.create(id=user_id, user_id=user_id)
            products_for_basket = Product.objects.filter(to_basket=True)

            for prod in products_for_basket:
                basket.products_to_basket.add(prod)
            basket.save()
            user_basket_serialized = BasketSerializer(basket).data
            return Response(user_basket_serialized)

    def delete(self, request, pk):
        user_id = request.user.id
        basket = Basket.objects.get(id=user_id)
        product = Product.objects.get(id=pk)
        product.to_basket = 0
        product.save()
        basket.products_to_basket.remove(product)
        return Response(status=status.HTTP_204_NO_CONTENT)

class ExportBasketInExcel(APIView):
    def get(self,request):
        user_id = request.user.id
        basket = Basket.objects.get(id=user_id)
        products_for_basket = Product.objects.filter(to_basket=True)

        alignment = Alignment(horizontal='center',
                              vertical='center', )
        workbook = Workbook()
        ws = workbook.active
        fields = ['Артикул', 'Наименование', 'Цена за шт с НДС', 'Количество']
        column = 1
        row = 1
        for field in fields:
            cell = ws.cell(column=column, row=1, value=field)
            cell.alignment = alignment
            column += 1
        for product in products_for_basket:
            row += 1
            ws.cell(column=1, row=row, value=product.article)
            ws.cell(column=2, row=row, value=product.name)
            ws.cell(column=3, row=row, value=product.price)
            ws.cell(column=4, row=row, value=product.amount)

        response = HttpResponse(content_type='application/ms_excel')
        response['Content_Disposition'] = f'attachment; filename = Data.xlsx'
        workbook.save(response)
        return response